from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict


def generate_pdf_report(game_data: Dict, leaderboard: List[Dict]) -> BytesIO:
    """
    PDF riport generálása
    """
    buffer = BytesIO()
    
    question_count = game_data.get('question_count', 0)
    pagesize = landscape(A4) if question_count >= 6 else A4
    
    doc = SimpleDocTemplate(buffer, pagesize=pagesize)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=1
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#34495e'),
        spaceAfter=12
    )
    
    title = Paragraph(game_data.get('quiz_title', 'Kvíz Játék'), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    game_info = [
        ['Játék kód:', game_data.get('game_code', 'N/A')],
        ['Dátum:', game_data.get('date', datetime.now().strftime('%Y-%m-%d %H:%M'))],
        ['Játékosok száma:', str(len(leaderboard))],
        ['Kérdések száma:', str(game_data.get('question_count', 0))]
    ]
    
    info_table = Table(game_info, colWidths=[5*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#7f8c8d')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#2c3e50')),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 1*cm))
    
    leaderboard_title = Paragraph('Eredménylista', heading_style)
    elements.append(leaderboard_title)
    
    note_style = ParagraphStyle(
        'Note',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#7f8c8d'),
        spaceAfter=8
    )
    note = Paragraph('Megjegyzés: A "#" jel azt jelzi, hogy a játékos offline volt az adott kérdés közben.', note_style)
    elements.append(note)
    elements.append(Spacer(1, 0.3*cm))
    
    question_count = game_data.get('question_count', 0)
    accuracy_col_idx = 4
    
    table_data = [['Helyezés', 'Becenév', 'Össz.\nPont', 'Helyes\nválaszok', 'Pontosság\n(%)']]
    
    for i in range(question_count):
        table_data[0].append(f'{i+1}.\nkérdés')
    
    for idx, player in enumerate(leaderboard, 1):
        accuracy = (player['correct_answers'] / player['total_answers'] * 100) if player['total_answers'] > 0 else 0
        
        row = [
            str(idx),
            player['nickname'],
            str(player['score']),
            f"{player['correct_answers']}/{player['total_answers']}",
            f"{accuracy:.1f}%"
        ]
        
        question_scores = player.get('question_scores', {})
        for q_idx in range(question_count):
            if q_idx in question_scores:
                q_data = question_scores[q_idx]
                if not q_data.get('was_online', True):
                    row.append('#')
                else:
                    row.append(str(q_data.get('points', 0)))
            else:
                row.append('-')
        
        table_data.append(row)
    
    col_widths = [1.8*cm, 4*cm, 2*cm, 2.5*cm, 2.2*cm] + [1.5*cm] * question_count
    
    leaderboard_table = Table(table_data, colWidths=col_widths)
    
    table_style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]
    
    # Első három helyezett kiemelése
    if len(leaderboard) >= 1:
        table_style_commands.append(('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ffd700')))
    if len(leaderboard) >= 2:
        table_style_commands.append(('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#c0c0c0')))
    if len(leaderboard) >= 3:
        table_style_commands.append(('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#cd7f32')))
    
    if len(leaderboard) > 3:
        table_style_commands.append(('BACKGROUND', (0, 4), (-1, -1), colors.beige))
    
    table_style = TableStyle(table_style_commands)
    
    leaderboard_table.setStyle(table_style)
    elements.append(leaderboard_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return buffer


def generate_excel_report(game_data: Dict, leaderboard: List[Dict]) -> BytesIO:
    """
    Excel riport generálása
    """
    buffer = BytesIO()
    wb = Workbook()
    
    ws_summary = wb.active
    ws_summary.title = "Összefoglaló"
    
    ws_summary['A1'] = game_data.get('quiz_title', 'Kvíz Játék')
    ws_summary['A1'].font = Font(size=18, bold=True, color='2C3E50')
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A3'] = 'Játék kód:'
    ws_summary['B3'] = game_data.get('game_code', 'N/A')
    ws_summary['A4'] = 'Dátum:'
    ws_summary['B4'] = game_data.get('date', datetime.now().strftime('%Y-%m-%d %H:%M'))
    ws_summary['A5'] = 'Játékosok száma:'
    ws_summary['B5'] = len(leaderboard)
    ws_summary['A6'] = 'Kérdések száma:'
    ws_summary['B6'] = game_data.get('question_count', 0)
    
    ws_summary['A8'] = 'Megjegyzés:'
    ws_summary['A9'] = 'A "#" jel a ranglistában azt jelzi, hogy a játékos offline volt az adott kérdés teljes ideje alatt.'
    ws_summary['A9'].font = Font(italic=True, color='7F8C8D', size=10)
    ws_summary.merge_cells('A9:D9')
    
    for row in range(3, 7):
        ws_summary[f'A{row}'].font = Font(bold=True, color='7F8C8D')
        ws_summary[f'B{row}'].font = Font(color='2C3E50')
    
    ws_summary['A8'].font = Font(bold=True, color='7F8C8D')
    
    ws_leaderboard = wb.create_sheet(title="Ranglista")
    
    question_count = game_data.get('question_count', 0)
    headers = ['Helyezés', 'Becenév', 'Összesített Pontszám', 'Helyes válaszok', 'Pontosság (%)']
    
    for i in range(question_count):
        headers.append(f'{i+1}. kérdés pontszáma')
    
    ws_leaderboard.append(headers)
    
    header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_leaderboard.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    medal_fills = {
        1: PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid'),
        2: PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid'),
        3: PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid'),
    }
    
    offline_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
    
    for idx, player in enumerate(leaderboard, 1):
        accuracy = (player['correct_answers'] / player['total_answers'] * 100) if player['total_answers'] > 0 else 0
        
        row = [
            idx,
            player['nickname'],
            player['score'],
            player['correct_answers'],
            f"{accuracy:.1f}%"
        ]
        
        question_scores = player.get('question_scores', {})
        for q_idx in range(question_count):
            if q_idx in question_scores:
                q_data = question_scores[q_idx]
                if not q_data.get('was_online', True):
                    row.append('#')
                else:
                    row.append(q_data.get('points', 0))
            else:
                row.append('-')
        
        ws_leaderboard.append(row)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws_leaderboard.cell(row=idx + 1, column=col_num)
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if idx in medal_fills and cell.value != '#':
                cell.fill = medal_fills[idx]
                cell.font = Font(bold=True)
            
            if cell.value == '#':
                cell.fill = offline_fill
                cell.font = Font(bold=True, color='CC0000')
    
    column_letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    
    ws_leaderboard.column_dimensions['A'].width = 12  # Helyezés
    ws_leaderboard.column_dimensions['B'].width = 20  # Becenév
    ws_leaderboard.column_dimensions['C'].width = 18  # Összesített Pontszám
    ws_leaderboard.column_dimensions['D'].width = 15  # Helyes válaszok
    ws_leaderboard.column_dimensions['E'].width = 15  # Pontosság
    
    for i in range(question_count):
        col_letter = column_letters[5 + i]
        ws_leaderboard.column_dimensions[col_letter].width = 12
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 25
    
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
